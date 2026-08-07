import streamlit as st
import pandas as pd
import io
import random
import re
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

# CONFIGURAÇÃO DA PÁGINA
st.set_page_config(page_title="AJENS - ROTEIROS DE TRANSPORTE", layout="wide", page_icon="🚌")
st.title("GERAR ROTEIROS - AJENS")

# CONFIGURAÇÃO DA FROTA PADRÃO
st.header("CONFIGURAÇÃO DA FROTA DE HOJE")
st.markdown("Configure cada veículo individualmente abaixo. Você pode **adicionar novas linhas**, **excluir veículos**, alterar o número de **vagas** e preencher o **nome do motorista**.")

if 'frota_padrao' not in st.session_state:
    st.session_state.frota_padrao = pd.DataFrame([
        {"Veículo": "Ônibus", "Vagas": 46, "Horário": "21:30", "Motorista": "Cláudio"},
        {"Veículo": "Micro-ônibus", "Vagas": 32, "Horário": "21:30", "Motorista": "Van"},
        {"Veículo": "Micro-ônibus", "Vagas": 32, "Horário": "21:30", "Motorista": "Whiray"},
        {"Veículo": "Micro-ônibus", "Vagas": 30, "Horário": "21:30", "Motorista": "Romário"},
        {"Veículo": "Micro-ônibus", "Vagas": 28, "Horário": "21:30", "Motorista": "George"},
        {"Veículo": "Van", "Vagas": 15, "Horário": "22:00", "Motorista": "Anselmo"},
        {"Veículo": "Van", "Vagas": 15, "Horário": "22:00", "Motorista": "Daniel"},
        {"Veículo": "Van", "Vagas": 15, "Horário": "22:00", "Motorista": "Higor"}
    ])

frota_df = st.data_editor(
    st.session_state.frota_padrao,
    column_config={
        "Veículo": st.column_config.SelectboxColumn("Tipo de Veículo", options=["Ônibus", "Micro-ônibus", "Van"], required=True),
        "Vagas": st.column_config.NumberColumn("Vagas Disponíveis", min_value=1, required=True),
        "Horário": st.column_config.SelectboxColumn("Horário de Retorno", options=["21:30", "22:00"], required=True),
        "Motorista": st.column_config.TextColumn("Nome do Motorista (Opcional)")
    },
    num_rows="dynamic",
    use_container_width=True,
    hide_index=True
)

# FUNÇÃO DE PRIORIDADE DE VEÍCULOS
def prioridade_veiculo(tipo):
    prio = {"Ônibus": 1, "Micro-ônibus": 2}
    return prio.get(tipo, 3)

frota_config = []
for v in frota_df.to_dict('records'):
    mot = str(v['Motorista']).strip() if pd.notna(v['Motorista']) and str(v['Motorista']).upper() != "NAN" else ""
    frota_config.append({
        'tipo': v['Veículo'],
        'cap': int(v['Vagas']) if pd.notna(v['Vagas']) else 0,
        'horario': v['Horário'],
        'motorista': mot
    })

frota_2130 = sorted([v for v in frota_config if v['horario'] == '21:30'], key=lambda x: prioridade_veiculo(x['tipo']))
frota_normal = sorted([v for v in frota_config if v['horario'] == '22:00'], key=lambda x: prioridade_veiculo(x['tipo']))

st.divider()

# UPLOAD DA PLANILHA
st.header("ANALISAR PLANILHA")
st.caption("Envie o arquivo exportado das respostas do formulário Google (.xlsx ou .csv).")
arquivo = st.file_uploader("Envie o ficheiro Excel ou CSV de respostas", type=['xlsx', 'csv'])

if arquivo:
    df = pd.read_csv(arquivo) if arquivo.name.endswith('.csv') else pd.read_excel(arquivo)
    df = df.dropna(how='all')
    
    # CRIA O CARIMBO DE FILA COM BASE NA ORDEM DO FORMULÁRIO
    df['ordem_preenchimento'] = range(len(df))

    colunas_lista = list(df.columns)

    def encontrar_indice(palavras_chave, indice_padrao):
        for i, col in enumerate(colunas_lista):
            if any(palavra in str(col).upper() for palavra in palavras_chave):
                return i
        return indice_padrao if indice_padrao < len(colunas_lista) else 0

    # SIDEBAR DE AJUSTES DE COLUNAS
    st.sidebar.header("AJUSTE DE COLUNAS")
    with st.sidebar.expander("Mapeamento", expanded=False):
        col_hora = st.selectbox("Data/Hora", colunas_lista, index=encontrar_indice(["CARIMBO", "HORA"], 0))
        col_nome = st.selectbox("Nome", colunas_lista, index=encontrar_indice(["NOME"], 3))
        col_inst = st.selectbox("Instituição", colunas_lista, index=encontrar_indice(["INSTITUIÇÃO"], 5))
        col_uso = st.selectbox("Uso (Ida/Volta)", colunas_lista, index=6 if len(colunas_lista) > 6 else 0)
        col_retorno = st.selectbox("Horário", colunas_lista, index=encontrar_indice(["RETORNO", "HORÁRIO"], 7))
        col_desembarque = st.selectbox("Bairro/Desembarque", colunas_lista, index=encontrar_indice(["BAIRRO", "DESEMBARQUE"], 8))

    # LIMPEZA E PADRONIZAÇÃO DE DADOS
    cols_essenciais = [col_nome, col_inst, col_uso, col_retorno, col_desembarque]
    for c in cols_essenciais:
        df[c] = df[c].fillna("").astype(str).str.strip()

    invalidos = ["NAN", "NONE", "NULL", ""]
    df = df[~df[col_nome].str.upper().isin(invalidos) & ~df[col_inst].str.upper().isin(invalidos)]
    df = df[~df[col_nome].str.contains("Não vou|Não vai|cancelada", case=False, na=False)]

    # FORÇA AS INSTITUIÇÕES EM MAIÚSCULO E LIMPA CARACTERES
    df[col_inst] = df[col_inst].str.replace("📚", "", regex=False).str.strip().str.upper()
    df[col_hora] = df[col_hora].astype(str)

    # NOVAS REGRAS DE INSTITUIÇÃO MISTA (DIVIDE EM SÓ IDA E SÓ RETORNO)
    novas_linhas = []
    indices_remover = []

    for idx, row in df.iterrows():
        inst_clean = " ".join(str(row[col_inst]).split())
        
        if "IDA UNEX M1 E RETORNO UNEX M3" in inst_clean:
            row_ida = row.copy()
            row_ida[col_inst] = "UNEX M1 SÃO JUDAS"
            row_ida[col_uso] = "SÓ IDA"
            
            row_volta = row.copy()
            row_volta[col_inst] = "UNEX M3 CÉSAR BORGES"
            row_volta[col_uso] = "SÓ RETORNO"
            
            novas_linhas.extend([row_ida, row_volta])
            indices_remover.append(idx)
            
        elif "IDA UNEX M3 E RETORNO UNEX M1" in inst_clean:
            row_ida = row.copy()
            row_ida[col_inst] = "UNEX M3 CÉSAR BORGES"
            row_ida[col_uso] = "SÓ IDA"
            
            row_volta = row.copy()
            row_volta[col_inst] = "UNEX M1 SÃO JUDAS"
            row_volta[col_uso] = "SÓ RETORNO"
            
            novas_linhas.extend([row_ida, row_volta])
            indices_remover.append(idx)
            
        elif "IDA MEDICINA E RETORNO M1" in inst_clean:
            row_ida = row.copy()
            row_ida[col_inst] = "UNEX MEDICINA"
            row_ida[col_uso] = "SÓ IDA"
            
            row_volta = row.copy()
            row_volta[col_inst] = "UNEX M1 SÃO JUDAS"
            row_volta[col_uso] = "SÓ RETORNO"
            
            novas_linhas.extend([row_ida, row_volta])
            indices_remover.append(idx)

    if indices_remover:
        df = df.drop(indices_remover)
        if novas_linhas:
            df = pd.concat([df, pd.DataFrame(novas_linhas)], ignore_index=True)

    # REGRAS DE ALOCAÇÃO
    def identificar_polo(nome):
        if "ANHANGUERA-SEDE" in nome: return "ANHANGUERA-SEDE"
        if "ANHANGUERA-CAP" in nome: return "ANHANGUERA-CAP"
        if any(x in nome for x in ["EXPERT", "EVOLUA"]): return "EXPERT/EVOLUA"
        if "HUMBERTO" in nome: return "HUMBERTO RIBEIRO"
        if "UESB" in nome: return "UESB"
        if any(x in nome for x in ["UNEX M1", "SÃO JUDAS"]): return "M1"
        if any(x in nome for x in ["UNEX M3", "CÉSAR BORGES"]): return "M3"
        if "MEDICINA" in nome: return "UNEX MEDICINA"
        if "UNIGRANDE" in nome: return "UNIGRANDE"
        if "UFSB" in nome: return "UFSB"
        if "IFBA" in nome: return "IFBA"
        if "FATESB" in nome: return "FATESB"
        if "PROCURSO" in nome: return "PROCURSO"

    # MATRIZ DE MISTURA
    matriz_mistura = {
        "M1": ["UNEX MEDICINA", "FATESB", "UNIGRANDE", "EXPERT/EVOLUA", "PROCURSO", "M3", "ANHANGUERA-SEDE", "ANHANGUERA-CAP"],
        "UNEX MEDICINA": ["M1", "FATESB", "UNIGRANDE", "EXPERT/EVOLUA", "PROCURSO", "M3", "ANHANGUERA-SEDE"],
        "M3": ["PROCURSO", "EXPERT/EVOLUA", "FATESB", "UNIGRANDE", "UESB", "M1", "ANHANGUERA-SEDE", "ANHANGUERA-CAP"],
        "ANHANGUERA-SEDE": ["ANHANGUERA-CAP", "UFSB", "IFBA", "UNIGRANDE", "EXPERT/EVOLUA", "PROCURSO", "M3", "HUMBERTO RIBEIRO", "UESB", "M1"],
        "ANHANGUERA-CAP": ["ANHANGUERA-SEDE", "UFSB", "IFBA", "UNIGRANDE", "EXPERT/EVOLUA", "PROCURSO", "M3", "HUMBERTO RIBEIRO", "UESB", "M1"],
        "EXPERT/EVOLUA": ["PROCURSO", "FATESB", "UNIGRANDE", "M3", "HUMBERTO RIBEIRO", "UESB", "M1", "UNEX MEDICINA", "ANHANGUERA-SEDE"],
        "PROCURSO": ["EXPERT/EVOLUA", "FATESB", "UNIGRANDE", "M3", "HUMBERTO RIBEIRO", "UESB", "M1", "UNEX MEDICINA", "ANHANGUERA-SEDE"],
        "UNIGRANDE": ["FATESB", "EXPERT/EVOLUA", "PROCURSO", "M3", "M1", "HUMBERTO RIBEIRO", "UESB", "ANHANGUERA-SEDE"],
        "FATESB": ["UNIGRANDE", "EXPERT/EVOLUA", "PROCURSO", "M3", "M1", "UNEX MEDICINA", "HUMBERTO RIBEIRO", "UESB", "ANHANGUERA-SEDE"],
        "UESB": ["HUMBERTO RIBEIRO", "M3", "EXPERT/EVOLUA", "PROCURSO", "FATESB", "UNIGRANDE", "ANHANGUERA-SEDE", "ANHANGUERA-CAP", "M1"],
        "HUMBERTO RIBEIRO": ["UESB", "M3", "EXPERT/EVOLUA", "PROCURSO", "UNIGRANDE", "ANHANGUERA-SEDE", "ANHANGUERA-CAP", "M1"],
        "IFBA": ["ANHANGUERA-SEDE", "ANHANGUERA-CAP", "UFSB", "UNIGRANDE", "FATESB", "EXPERT/EVOLUA", "PROCURSO", "M3", "HUMBERTO RIBEIRO", "UESB", "M1"],
        "UFSB": ["ANHANGUERA-SEDE", "ANHANGUERA-CAP", "IFBA", "HUMBERTO RIBEIRO", "UESB", "M3", "EXPERT/EVOLUA", "M1"]
    }

    def classificar_uso(uso_str):
        u = str(uso_str).upper()
        tem_ida, tem_volta = "IDA" in u, "RETORNO" in u or "VOLTA" in u
        if tem_ida and tem_volta: return (0, "IDA E RETORNO")
        if tem_volta: return (1, "SÓ RETORNO")
        if tem_ida: return (2, "SÓ IDA")
        return (3, "OUTRO")

    def calcular_ocupacao(alunos_lista):
        inst_uso = defaultdict(lambda: {'ida': 0, 'volta': 0, 'outros': 0})
        for a in alunos_lista:
            _, uso_desc = classificar_uso(a.get(col_uso, ''))
            inst = a.get(col_inst, '')
            if uso_desc == "SÓ IDA":
                inst_uso[inst]['ida'] += 1
            elif uso_desc == "SÓ RETORNO":
                inst_uso[inst]['volta'] += 1
            else:
                inst_uso[inst]['outros'] += 1
                
        ocupacao_total = 0
        for contagem in inst_uso.values():
            pares = min(contagem['ida'], contagem['volta'])
            sobras_ida = contagem['ida'] - pares
            sobras_volta = contagem['volta'] - pares
            ocupacao_total += pares + sobras_ida + sobras_volta + contagem['outros']
        return ocupacao_total

    def alocar_frota_inteligente(alunos_lista, frota_disponivel):
        veiculos_utilizados = []
        alunos_restantes = list(alunos_lista)

        pref_onibus = ["M1", "M3", "ANHANGUERA-SEDE", "ANHANGUERA-CAP", "UNIGRANDE", "PROCURSO", "MEDICINA", "OUTROS"]
        pref_micros = ["M1", "M3", "ANHANGUERA-SEDE","UESB", "ANHANGUERA-CAP", "UNIGRANDE", "HUMBERTO RIBEIRO", "MEDICINA", "EXPERT/EVOLUA", "PROCURSO", "FATESB", "IFBA", "UFSB", "OUTROS"]

        for i, veiculo in enumerate(frota_disponivel):
            if not alunos_restantes: break

            veiculo_atual = {
                'id': f"{veiculo['tipo']} {i+1}",
                'tipo': veiculo['tipo'],
                'cap': veiculo['cap'],
                'horario': veiculo['horario'],
                'motorista': veiculo['motorista'],
                'alunos': []
            }

            demanda_polos = defaultdict(int)
            for a in alunos_restantes:
                demanda_polos[identificar_polo(a[col_inst])] += 1

            ancora = None
            lista_prefs = pref_onibus if veiculo_atual['tipo'] == 'Ônibus' else pref_micros
            
            lista_prefs_ordenada = sorted(
                [p for p in lista_prefs if demanda_polos[p] > 0],
                key=lambda p: (demanda_polos[p], -lista_prefs.index(p)),
                reverse=True
            )

            for polo in lista_prefs_ordenada:
                if veiculo_atual['tipo'] == 'Ônibus' and polo == 'UNIGRANDE':
                    if any(identificar_polo(a[col_inst]) != 'UNIGRANDE' for a in alunos_restantes):
                        continue
                ancora = polo
                break

            if not ancora and alunos_restantes: 
                ancora = identificar_polo(alunos_restantes[0][col_inst])

            polos_para_puxar = [ancora] + matriz_mistura.get(ancora, [])

            if veiculo_atual['horario'] == '22:00':
                tem_medicina = (ancora == "MEDICINA") or any(identificar_polo(a[col_inst]) == "MEDICINA" for a in veiculo_atual['alunos'])
                if tem_medicina:
                    polos_para_puxar = [p for p in polos_para_puxar if p not in ["UESB", "M3"]]
                if ancora in ["UESB", "M3"]:
                    polos_para_puxar = [p for p in polos_para_puxar if p != "MEDICINA"]

            polos_para_puxar = sorted(
                list(dict.fromkeys(polos_para_puxar)), 
                key=lambda p: (demanda_polos.get(p, 0), p == ancora), 
                reverse=True
            )

            for polo_atual in polos_para_puxar:
                if demanda_polos.get(polo_atual, 0) == 0:
                    continue
                    
                alunos_do_polo = [a for a in alunos_restantes if identificar_polo(a[col_inst]) == polo_atual]

                for aluno in alunos_do_polo:
                    if calcular_ocupacao(veiculo_atual['alunos'] + [aluno]) <= veiculo_atual['cap']:
                        veiculo_atual['alunos'].append(aluno)
                        alunos_restantes.remove(aluno)

            if veiculo_atual['alunos']:
                veiculos_utilizados.append(veiculo_atual)

        return veiculos_utilizados, alunos_restantes

    # PREPARAÇÃO INICIAL DAS FILAS
    df_2130 = df[df[col_retorno].str.contains("21:30", case=False, na=False)].copy()
    df_normal = df[~df[col_retorno].str.contains("21:30", case=False, na=False)].copy()

    # REGRA: Forçar Medicina para 22h (exceto se for SÓ IDA ou se a Instituição tiver "M1" no nome)
    is_medicina_2130 = df_2130[col_inst].str.contains("MEDICINA", case=False, na=False)
    is_so_ida = df_2130[col_uso].str.contains("SÓ IDA|SO IDA", case=False, na=False)
    is_excecao = df_2130[col_inst].str.contains("M1", case=False, na=False) | is_so_ida
    
    invalids = df_2130[is_medicina_2130 & ~is_excecao]
    df_2130 = df_2130.drop(invalids.index)
    df_normal = pd.concat([df_normal, invalids])

    # ORDENA AMBAS AS LISTAS ESTRITAMENTE PELA ORDEM DE PREENCHIMENTO DO FORMULÁRIO
    df_2130 = df_2130.sort_values(by='ordem_preenchimento')
    df_normal = df_normal.sort_values(by='ordem_preenchimento')

    vagas_2130_total = sum(v['cap'] for v in frota_2130)
    alunos_2130 = df_2130.to_dict('records')

    # PASSO 1: ANÁLISE GLOBAL E DEFINIÇÃO DA LISTA DE ESPERA (ANTES DE DIVIDIR CARROS)
    contemplados_2130 = []
    lista_espera_2130 = []

    for aluno in alunos_2130:
        if calcular_ocupacao(contemplados_2130 + [aluno]) <= vagas_2130_total:
            contemplados_2130.append(aluno)
        else:
            aluno['remanejado_2130'] = True
            lista_espera_2130.append(aluno)

    # DIVIDIR OS CARROS DE 21H30 EXCLUSIVAMENTE COM OS CONTEMPLADOS
    roteiros_2130, sobra_roteiro_2130 = alocar_frota_inteligente(contemplados_2130, frota_2130)

    if sobra_roteiro_2130 and roteiros_2130:
        for aluno in sobra_roteiro_2130:
            roteiros_2130[-1]['alunos'].append(aluno)
    elif sobra_roteiro_2130 and not roteiros_2130:
        for aluno in sobra_roteiro_2130:
            aluno['remanejado_2130'] = True
            lista_espera_2130.append(aluno)

    # INTEGRAR A LISTA DE ESPERA DIRETAMENTE NO ROTEIRO DE 22H00
    alunos_normal = lista_espera_2130 + df_normal.to_dict('records')
    roteiros_normal, alunos_sem_vaga = alocar_frota_inteligente(alunos_normal, frota_normal)

    roteiros_prontos = roteiros_2130 + roteiros_normal

    for rot in roteiros_prontos:
        if rot.get('alunos'):
            rot['monitor'] = random.choice(rot['alunos'])

    # SISTEMA DE CORES
    PALETA_CORES = [
        "FF66FF", "92D050", "FFC896", "9DC3E6", "C6E0B4",
        "F4B183", "B4A7D6", "8EA9DB", "F8CBAD", "BFBFBF",
        "E6B8AF", "C5E1A5", "FFE082", "B39DDB", "80CBC4"
    ]
    COR_PREDEFINIDA = {
        "UNIGRANDE": "FF66FF", "ANHANGUERA": "92D050", "M1": "FFC896",
        "M3": "9DC3E6", "UESB": "C6E0B4", "EXPERT": "F4B183",
        "HUMBERTO": "B4A7D6", "MEDICINA": "80CBC4", "OUTROS": "BFBFBF"
    }

    mapa_cores_instituicoes = {}
    cores_usadas = set()

    def obter_cor(inst_nome):
        if inst_nome in mapa_cores_instituicoes:
            return mapa_cores_instituicoes[inst_nome]
        
        for chave, cor in COR_PREDEFINIDA.items():
            if chave in inst_nome:
                mapa_cores_instituicoes[inst_nome] = cor
                cores_usadas.add(cor)
                return cor

        for cor in PALETA_CORES:
            if cor not in cores_usadas:
                mapa_cores_instituicoes[inst_nome] = cor
                cores_usadas.add(cor)
                return cor

        mapa_cores_instituicoes[inst_nome] = "D9D9D9"
        return "D9D9D9"

    for rot in roteiros_prontos:
        for aluno in rot['alunos']:
            obter_cor(str(aluno[col_inst]))

    # RESUMO DA OPERAÇÃO
    st.subheader("RESUMO DA OPERAÇÃO")
    vagas_2130_total = sum(v['cap'] for v in frota_2130)
    demanda_2130 = len(df_2130)
    saldo_2130 = vagas_2130_total - demanda_2130

    vagas_normal_total = sum(v['cap'] for v in frota_normal)
    demanda_normal = len(df_normal) + max(0, demanda_2130 - vagas_2130_total)
    saldo_normal = vagas_normal_total - demanda_normal

    c1, c2, c3 = st.columns(3)
    c1.metric("Vagas 21:30h", vagas_2130_total)
    c2.metric("Demanda 21:30h", demanda_2130)
    c3.metric("Saldo 21:30h", f"{saldo_2130} vagas")

    c4, c5, c6 = st.columns(3)
    c4.metric("Vagas 22:00h", vagas_normal_total)
    c5.metric("Demanda 22:00h", demanda_normal)
    c6.metric("Saldo 22:00h", f"{saldo_normal} vagas")

    if alunos_sem_vaga:
        st.error(f"⚠️ ATENÇÃO: Faltou espaço para {len(alunos_sem_vaga)} aluno(s). Adicione um veículo na tabela acima.")
    else:
        st.success("✅ Todos os estudantes foram alocados com sucesso!")

    lista_espera_2130 = [aluno for aluno in alunos_2130 if aluno.get('remanejado_2130', False)]
    if lista_espera_2130:
        lista_espera_df = pd.DataFrame(lista_espera_2130)
        lista_espera_df = lista_espera_df[[col_hora, col_nome, col_inst, col_uso, col_desembarque]].copy()
        lista_espera_df.columns = ['Data/Hora', 'Nome do Associado', 'Instituição', 'Uso do Transporte', 'Bairro']
        lista_espera_df['Bairro'] = lista_espera_df['Bairro'].apply(lambda x: "" if str(x).upper() in invalidos else x)
        lista_espera_df = lista_espera_df.reset_index(drop=True)
        st.subheader("🕒 Lista de Espera - Vagas 21:30h")
        st.caption("Estes estudantes ficaram fora das vagas das 21:30h e foram remanejados para os carros das 22:00h, na ordem do preenchimento do formulário.")
        st.dataframe(lista_espera_df, use_container_width=True)

    # EXIBIÇÃO VISUAL DOS ROTEIROS
    st.subheader("📋 ROTEIROS PRONTOS")
    col1, col2 = st.columns(2)

    for idx, rot in enumerate(roteiros_prontos):
        coluna_atual = col1 if idx % 2 == 0 else col2
        
        usos = [classificar_uso(a.get(col_uso, ''))[1] for a in rot['alunos']]
        t_ida = usos.count("IDA E RETORNO") + usos.count("SÓ IDA")
        t_volta = usos.count("IDA E RETORNO") + usos.count("SÓ RETORNO")
        
        ocupacao_real = calcular_ocupacao(rot['alunos'])

        mot = str(rot.get('motorista', '')).strip().upper()
        mot_texto = mot if mot and mot != "NAN" else "___________________________"

        insts = list(dict.fromkeys([str(a[col_inst]) for a in rot['alunos']]))

        with coluna_atual:
            with st.container(border=True):
                st.markdown(f"**{rot['id'].upper()}: MOTORISTA {mot_texto}**")
                st.markdown(f"**Vagas Ocupadas: {ocupacao_real}/{rot['cap']} (Pessoas: {len(rot['alunos'])}) | {t_ida} IDA, {t_volta} VOLTA**")

                for instituicao in insts:
                    cor_hex = mapa_cores_instituicoes.get(instituicao, "D9D9D9")
                    alunos_inst = [a for a in rot['alunos'] if str(a[col_inst]) == instituicao]
                    alunos_ord = [a for a in alunos_inst if not a.get('remanejado_2130', False)] + [a for a in alunos_inst if a.get('remanejado_2130', False)]
                    
                    st.markdown(
                        f"<div style='background-color:#{cor_hex};color:black;font-weight:bold;"
                        f"padding:6px;text-align:center;border-radius:4px;margin:8px 0 2px 0;'>"
                        f"{instituicao} {rot['horario']}H</div>",
                        unsafe_allow_html=True
                    )

                    df_exib = pd.DataFrame(alunos_ord)[[col_hora, col_nome, col_uso, col_desembarque]].copy()
                    df_exib.columns = ['Data/Hora', 'Nome do Associado', 'Uso do Transporte', 'Bairro']
                    
                    df_exib['Bairro'] = df_exib['Bairro'].apply(lambda x: "" if str(x).upper() in invalidos else x)
                    df_exib.loc[df_exib['Uso do Transporte'].str.upper().str.contains("IDA") & ~df_exib['Uso do Transporte'].str.upper().str.contains("RETORNO|VOLTA"), 'Bairro'] = "NÃO RETORNO"
                    df_exib['__remanejado_2130__'] = [bool(a.get('remanejado_2130', False)) for a in alunos_ord]
                    df_exib.index = range(1, len(df_exib) + 1)

                    monitor_nome = str((rot.get('monitor') or {}).get(col_nome, '')).strip().upper()

                    def destacar_linhas(row):
                        styles = [''] * len(row)
                        aluno_nome = str(row['Nome do Associado']).strip().upper()
                        grupo = classificar_uso(row['Uso do Transporte'])[1]
                        remanejado = bool(row['__remanejado_2130__'])
                        
                        if monitor_nome and aluno_nome == monitor_nome:
                            styles = ['background-color: #2E7D32; color: #FFFFFF; font-weight: bold'] * len(row)
                        elif remanejado:
                            # SINALIZAÇÃO EM AZUL PARA ALUNOS REMANEJADOS DAS 21H30
                            if grupo in {"SÓ RETORNO", "SÓ IDA"}:
                                styles = ['background-color: #FFFF00; color: #2563EB; font-weight: bold'] * len(row)
                            else:
                                styles = ['color: #2563EB; font-weight: bold'] * len(row)
                        elif grupo in {"SÓ RETORNO", "SÓ IDA"}:
                            styles = ['background-color: #FFFF00; color: #000'] * len(row)

                        return styles

                    styler = df_exib.style.apply(destacar_linhas, axis=1).hide(subset=['__remanejado_2130__'], axis='columns')
                    st.dataframe(styler, use_container_width=True)

    # EXPORTAÇÃO PARA EXCEL
    st.divider()
    st.subheader("BAIXAR ROTEIROS EM PLANILHA EXCEL")

    FILL_AMARELO = PatternFill("solid", fgColor="FFFF00")
    FILL_VERDE = PatternFill("solid", fgColor="2E7D32")
    FONTE_TITULO = Font(bold=True, size=18)
    FONTE_TOTAL = Font(bold=True, size=12)
    BORDA = Border(left=Side(style='thin', color='B7B7B7'), right=Side(style='thin', color='B7B7B7'),
                   top=Side(style='thin', color='B7B7B7'), bottom=Side(style='thin', color='B7B7B7'))

    def escrever_roteiro_excel(ws, linha, rot):
        mot = str(rot.get('motorista', '')).strip().upper()
        mot_texto = mot if mot and mot != "NAN" else "_______________"
        
        usos = [classificar_uso(a.get(col_uso, ''))[1] for a in rot['alunos']]
        qtd_ida = usos.count("IDA E RETORNO") + usos.count("SÓ IDA")
        qtd_volta = usos.count("IDA E RETORNO") + usos.count("SÓ RETORNO")

        cel = ws.cell(row=linha, column=1, value=f"{rot['id'].upper()}: MOTORISTA {mot_texto}")
        cel.font = FONTE_TITULO
        cel.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=4)
        linha += 1

        insts = list(dict.fromkeys([str(a[col_inst]) for a in rot['alunos']]))

        for inst in insts:
            cor_hex = mapa_cores_instituicoes.get(inst, "D9D9D9")
            fill_inst = PatternFill("solid", fgColor=cor_hex)

            cel_b = ws.cell(row=linha, column=1, value=f"{inst} {rot['horario']}H")
            cel_b.font = FONTE_TITULO
            cel_b.alignment = Alignment(horizontal='center', vertical='center')
            
            for c in range(1, 5):
                ws.cell(row=linha, column=c).fill = fill_inst
            ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=4)
            linha += 1

            alunos_ord = [a for a in rot['alunos'] if str(a[col_inst]) == inst and not a.get('remanejado_2130', False)] + [a for a in rot['alunos'] if str(a[col_inst]) == inst and a.get('remanejado_2130', False)]
            monitor_nome = str((rot.get('monitor') or {}).get(col_nome, '')).strip().upper()

            for aluno in alunos_ord:
                grupo = classificar_uso(aluno.get(col_uso, ''))[1]
                aluno_nome = str(aluno.get(col_nome, '')).strip().upper()
                remanejado = bool(aluno.get('remanejado_2130', False))
                
                bairro = str(aluno.get(col_desembarque, '')).strip()
                bairro_val = "" if bairro.upper() in invalidos else bairro
                if grupo == "SÓ IDA": bairro_val = "NÃO RETORNO"
                
                vals = [str(aluno.get(col_hora, '')), aluno_nome, str(aluno.get(col_uso, '')), bairro_val]
                
                for c, val in enumerate(vals, start=1):
                    cel_c = ws.cell(row=linha, column=c, value=val)
                    cel_c.border = BORDA
                    if monitor_nome and aluno_nome == monitor_nome:
                        cel_c.fill = FILL_VERDE
                        cel_c.font = Font(color="FFFFFF", bold=True)
                    elif remanejado:
                        # PLANILHA EXCEL: TEXTO EM AZUL PARA OS REMANEJADOS DAS 21H30
                        cel_c.font = Font(color="2563EB", bold=True)
                    elif grupo in {"SÓ RETORNO", "SÓ IDA"}:
                        cel_c.fill = FILL_AMARELO
                linha += 1

            linha += 1

        ws.cell(row=linha, column=1, value=f"TOTAL: IDA {qtd_ida}, RETORNO {qtd_volta}")
        cel_total = ws.cell(row=linha, column=1)
        cel_total.font = Font(bold=True, size=18)
        cel_total.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=4)
        return linha + 3

    if roteiros_prontos:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            wb = writer.book
            ws = wb.create_sheet("ROTEIRO DE IDA")
            
            linha_atual = 1
            for rot in roteiros_prontos:
                linha_atual = escrever_roteiro_excel(ws, linha_atual, rot)

            for col, w in zip(['A', 'B', 'C', 'D'], [20, 32, 20, 24]):
                ws.column_dimensions[col].width = w

            if lista_espera_2130:
                ws_realocados = wb.create_sheet("LISTA DE ESPERA 21h30")
                ws_realocados.cell(row=1, column=1, value="Data/Hora")
                ws_realocados.cell(row=1, column=2, value="Nome do Associado")
                ws_realocados.cell(row=1, column=3, value="Instituição")
                ws_realocados.cell(row=1, column=4, value="Uso do Transporte")
                ws_realocados.cell(row=1, column=5, value="Bairro")
                for idx, row in enumerate(lista_espera_df.itertuples(index=False, name=None), start=2):
                    ws_realocados.cell(row=idx, column=1, value=row[0])
                    ws_realocados.cell(row=idx, column=2, value=row[1])
                    ws_realocados.cell(row=idx, column=3, value=row[2])
                    ws_realocados.cell(row=idx, column=4, value=row[3])
                    ws_realocados.cell(row=idx, column=5, value=row[4])
                for col, w in zip(['A', 'B', 'C', 'D', 'E'], [20, 32, 24, 20, 24]):
                    ws_realocados.column_dimensions[col].width = w

        hoje = datetime.now()
        nome_arquivo = f"{hoje.day:02d}-{hoje.month:02d}-ROTEIRO-{hoje.year}.xlsx"

        st.download_button(
            label="BAIXAR PLANILHA DE ROTEIROS (.xlsx)",
            data=output.getvalue(),
            file_name=nome_arquivo,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )